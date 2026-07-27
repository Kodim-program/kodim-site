import re

import openpyxl
from django import forms
from django.contrib import admin, messages
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import path, reverse
from django.utils.html import format_html
from django_ckeditor_5.widgets import CKEditor5Widget

from .forms import QuestionImportForm
from .models import (
    Material, Enrollment, MaterialProgress,
    Test, Question, Answer, TestResult, TestResultAnswer,
)

# Значення в колонці question_type Excel-файлу (формат "пар"), які розпізнаються
# як питання "кілька правильних відповідей". Усе інше -> одна правильна.
MULTIPLE_TYPE_VALUES = {'кілька', 'декілька', 'multiple', 'кілька відповідей'}

# Значення в колонках correct_N (формат "пар"), які означають "правильна відповідь"
CORRECT_VALUES = {'так', 'yes', 'true', '1', '+', 'правильно'}

# Заголовки колонок для "буквеного" формату (Питання / A / B / C / D / Правильна),
# який зручний вчителям — кожна відповідь у своїй колонці, правильна вказана літерою.
QUESTION_HEADER_ALIASES = {'question', 'питання', 'текст питання', 'текст'}
ORDER_HEADER_ALIASES = {'order', '№', '#', 'номер', 'номер питання'}
CORRECT_LETTER_HEADER_ALIASES = {'correct', 'правильна', 'правильна відповідь', 'відповідь', 'answer'}


# ── Матеріали курсу ────────────────────────────────────────────
class MaterialAdminForm(forms.ModelForm):
    class Meta:
        model = Material
        fields = '__all__'
        widgets = {
            'content': CKEditor5Widget(config_name='news'),
        }


@admin.register(Material)
class MaterialAdmin(admin.ModelAdmin):
    form = MaterialAdminForm
    list_display = ('course', 'order', 'title')
    list_filter = ('course',)
    ordering = ('course', 'order')
    search_fields = ('title', 'course__name')


# ── Призначення курсів учням ───────────────────────────────────
@admin.register(Enrollment)
class EnrollmentAdmin(admin.ModelAdmin):
    list_display = ('user', 'course', 'assigned_at', 'progress_display')
    list_filter = ('course',)
    search_fields = ('user__username', 'user__email', 'course__name')
    autocomplete_fields = ('user',)

    @admin.display(description='Прогрес')
    def progress_display(self, obj):
        return f'{obj.progress_percent()}% ({obj.read_materials_count()}/{obj.total_materials()})'


@admin.register(MaterialProgress)
class MaterialProgressAdmin(admin.ModelAdmin):
    list_display = ('user', 'material', 'is_read', 'read_at')
    list_filter = ('is_read', 'material__course')
    search_fields = ('user__username',)


# ── Тести ───────────────────────────────────────────────────────
class AnswerInline(admin.TabularInline):
    model = Answer
    extra = 4


@admin.register(Question)
class QuestionAdmin(admin.ModelAdmin):
    list_display = ('test', 'order', 'text', 'question_type')
    list_filter = ('test',)
    inlines = [AnswerInline]


class QuestionInline(admin.TabularInline):
    model = Question
    extra = 1
    show_change_link = True


@admin.register(Test)
class TestAdmin(admin.ModelAdmin):
    list_display = ('material', 'title', 'passing_score', 'max_attempts', 'import_link')
    list_filter = ('material__course',)
    autocomplete_fields = ('material',)
    inlines = [QuestionInline]

    # ── Кнопка "Імпортувати з Excel" у списку тестів ──────────────
    @admin.display(description='Питання')
    def import_link(self, obj):
        url = reverse('admin:learning_test_import_questions', args=[obj.pk])
        return format_html(
            '<a class="button" href="{}">Імпортувати з Excel</a>', url,
        )

    # ── Додаткові URL-и для імпорту ────────────────────────────────
    def get_urls(self):
        custom_urls = [
            path(
                '<int:test_id>/import-questions/',
                self.admin_site.admin_view(self.import_questions_view),
                name='learning_test_import_questions',
            ),
            path(
                'import-template/',
                self.admin_site.admin_view(self.download_template_view),
                name='learning_test_import_template',
            ),
        ]
        return custom_urls + super().get_urls()

    def import_questions_view(self, request, test_id):
        test = get_object_or_404(Test, pk=test_id)

        if request.method == 'POST':
            form = QuestionImportForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    created_questions, created_answers = self._import_questions_from_excel(
                        test=test,
                        file_obj=form.cleaned_data['file'],
                        replace_existing=form.cleaned_data['replace_existing'],
                    )
                except ValueError as exc:
                    messages.error(request, str(exc))
                else:
                    messages.success(
                        request,
                        f'Імпортовано {created_questions} питань '
                        f'та {created_answers} варіантів відповідей у тест «{test}».',
                    )
                    return redirect('admin:learning_test_change', test.pk)
        else:
            form = QuestionImportForm()

        return render(request, 'admin/learning/test_import.html', {
            'form': form,
            'test': test,
            'opts': self.model._meta,
            'template_download_url': reverse('admin:learning_test_import_template'),
        })

    def download_template_view(self, request):
        """Готовий Excel-шаблон із прикладами, щоб вчитель бачив формат колонок."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Питання'
        sheet.append([
            'order', 'question_type', 'question',
            'answer_1', 'correct_1', 'answer_2', 'correct_2',
            'answer_3', 'correct_3', 'answer_4', 'correct_4',
        ])
        sheet.append([
            1, 'одна', 'Скільки буде 2 + 2?',
            '3', '', '4', 'так', '5', '', '', '',
        ])
        sheet.append([
            2, 'кілька', 'Які з цього — мови програмування?',
            'Python', 'так', 'HTML', '', 'Java', 'так', 'CSS', '',
        ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="shablon_pytan.xlsx"'
        workbook.save(response)
        return response

    def _import_questions_from_excel(self, test, file_obj, replace_existing):
        """
        Підтримує ДВА формати аркуша (визначається автоматично за заголовками):

        1) "Буквений" формат (зручний вчителям):
           № | Питання | A | B | C | D | Правильна
           Кожна відповідь у своїй колонці, правильна(-і) вказана(-і) літерою
           колонки у стовпці "Правильна" (можна декілька через кому/пробіл — тоді
           питання автоматично стане типу "кілька правильних відповідей").

        2) Формат "пар" (сумісність зі старим шаблоном):
           order | question_type | question | answer_1 | correct_1 | answer_2 | correct_2 | ...
        """
        try:
            workbook = openpyxl.load_workbook(file_obj, data_only=True)
        except Exception:
            raise ValueError(
                'Не вдалося прочитати файл. Перевірте, що це коректний .xlsx-файл.'
            )

        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise ValueError('Файл порожній.')

        headers = [str(h or '').strip().lower() for h in header_row]
        data_rows = list(rows_iter)

        question_col = self._find_header_index(headers, QUESTION_HEADER_ALIASES)
        correct_col = self._find_header_index(headers, CORRECT_LETTER_HEADER_ALIASES)

        if question_col is not None and correct_col is not None:
            return self._import_letter_format(
                test, headers, data_rows, question_col, correct_col, replace_existing,
            )
        return self._import_pairs_format(test, data_rows, replace_existing)

    @staticmethod
    def _find_header_index(headers, aliases):
        for i, header in enumerate(headers):
            if header in aliases:
                return i
        return None

    def _import_letter_format(self, test, headers, rows, question_col, correct_col, replace_existing):
        order_col = self._find_header_index(headers, ORDER_HEADER_ALIASES)
        option_cols = [
            i for i in range(len(headers)) if i not in (question_col, correct_col, order_col)
        ]
        # Мітка варіанту відповіді = текст заголовка колонки (A, B, В, Г ...),
        # або літера за позицією, якщо заголовок порожній.
        option_labels = {}
        for position, i in enumerate(option_cols):
            label = headers[i].strip().upper() if headers[i] else ''
            option_labels[i] = label or chr(ord('A') + position)

        if replace_existing:
            test.questions.all().delete()

        created_questions = 0
        created_answers = 0

        for row_number, row in enumerate(rows, start=1):
            if not row or row[question_col] in (None, ''):
                continue
            text = str(row[question_col]).strip()
            if not text:
                continue

            order = row_number
            if order_col is not None and row[order_col] not in (None, ''):
                try:
                    order = int(row[order_col])
                except (TypeError, ValueError):
                    pass

            correct_raw = str(row[correct_col] or '').strip().upper()
            if re.search(r'[,;/\s]', correct_raw):
                correct_letters = {p for p in re.split(r'[,;/\s]+', correct_raw) if p}
            else:
                correct_letters = set(correct_raw) if correct_raw else set()

            question_type = Question.MULTIPLE if len(correct_letters) > 1 else Question.SINGLE

            question = Question.objects.create(
                test=test, text=text, question_type=question_type, order=order,
            )
            created_questions += 1

            for i in option_cols:
                option_text = row[i]
                if option_text in (None, ''):
                    continue
                Answer.objects.create(
                    question=question,
                    text=str(option_text).strip(),
                    is_correct=option_labels[i] in correct_letters,
                )
                created_answers += 1

        if created_questions == 0:
            raise ValueError('У файлі не знайдено жодного питання для імпорту.')

        return created_questions, created_answers

    def _import_pairs_format(self, test, rows, replace_existing):
        if not rows:
            raise ValueError('Файл порожній або не містить рядків із питаннями (крім заголовка).')

        if replace_existing:
            test.questions.all().delete()

        created_questions = 0
        created_answers = 0

        for row in rows:
            if not row or row[0] in (None, ''):
                continue

            try:
                order = int(row[0])
            except (TypeError, ValueError):
                raise ValueError(f'Некоректний номер (order) у рядку: {row}')

            type_raw = str(row[1] or '').strip().lower()
            question_type = (
                Question.MULTIPLE if type_raw in MULTIPLE_TYPE_VALUES else Question.SINGLE
            )

            text = str(row[2] or '').strip()
            if not text:
                continue

            question = Question.objects.create(
                test=test, text=text, question_type=question_type, order=order,
            )
            created_questions += 1

            for i in range(3, len(row), 2):
                answer_text = row[i]
                if answer_text in (None, ''):
                    continue
                correct_raw = str(row[i + 1] or '').strip().lower() if i + 1 < len(row) else ''
                Answer.objects.create(
                    question=question,
                    text=str(answer_text).strip(),
                    is_correct=correct_raw in CORRECT_VALUES,
                )
                created_answers += 1

        return created_questions, created_answers


class TestResultAnswerInline(admin.TabularInline):
    """Показує вчителю, що саме обрав учень по кожному питанню цієї спроби."""
    model = TestResultAnswer
    extra = 0
    can_delete = False
    fields = ('question', 'selected_answers_display', 'is_correct')
    readonly_fields = ('question', 'selected_answers_display', 'is_correct')

    def has_add_permission(self, request, obj=None):
        return False

    @admin.display(description='Обрані варіанти')
    def selected_answers_display(self, obj):
        return ', '.join(a.text for a in obj.selected_answers.all()) or '—'


@admin.register(TestResult)
class TestResultAdmin(admin.ModelAdmin):
    list_display = ('user', 'test', 'score', 'passed', 'completed_at')
    list_filter = ('test', 'passed')
    search_fields = ('user__username',)
    inlines = [TestResultAnswerInline]
